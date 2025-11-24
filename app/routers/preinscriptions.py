# app/routers/preinscriptions.py
from __future__ import annotations

import base64
import json
import logging
import os
import re
import shutil
from datetime import datetime, timezone, date as _date
from pathlib import Path
from typing import Optional, Set

from fastapi import (
    APIRouter, Request, Depends, Form, HTTPException,
    Query, UploadFile, File
)
from fastapi.responses import HTMLResponse, RedirectResponse, FileResponse, Response
from sqlalchemy import func, text
from sqlmodel import Session, select

import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
import pandas as pd

from ..core.database import get_session
from ..core.middleware import get_shared_session
from ..core.program_schema_integration import (
    SchemaRoutingService, table_exists_anywhere, get_schema_routing_service,
    ProgramSchemaService, get_current_programme_from_session
)
from ..core.config import settings
from ..core.path_config import path_config
from ..services.file_upload_service import FileUploadService
from ..core.security import get_current_user
from ..templates import templates

from ..models.base import (
    Programme, Candidat, Entreprise,
    StatutDossier, Document
)
from ..models.preinscription import Preinscription, Eligibilite

# Enums
try:
    from ..models.enums import TypeDocument  # recommandé
except Exception:
    try:
        from ..models.base import TypeDocument  # fallback si défini là
    except Exception:
        TypeDocument = None  # pas d'enum dispo

from ..services.geocoding import geocode_one
from ..services.eligibilite import evaluate_eligibilite, entreprise_age_annees
from ..services.uploads import validate_upload  # limites taille/type

router = APIRouter()

# ---------- Constantes / helpers ----------
SAFE_RE = re.compile(r"[^A-Za-z0-9._-]+")

# Liste proposée au front (si l'enum existe on s'aligne)
if TypeDocument:
    DOC_TYPES_DEFAULT = [td.value for td in TypeDocument]
else:
    DOC_TYPES_DEFAULT = ["CNI", "KBIS", "JUSTIFICATIF_DOMICILE", "RIB", "CV", "DIPLOME", "ATTESTATION", "AUTRE"]


def safe_name(s: str) -> str:
    """Nettoie un titre de document pour le rendre filesystem-friendly."""
    s = (s or "").strip().replace(" ", "_")
    s = SAFE_RE.sub("_", s)
    return s[:120] or "doc"


def ensure_media_root() -> Path:
    """Récupère le MEDIA_ROOT (config) et s'assure qu'il existe."""
    from ..core.config import Settings
    settings = Settings()
    root = settings.MEDIA_ROOT
    return root


def save_upload(dst: Path, file: UploadFile):
    dst.parent.mkdir(parents=True, exist_ok=True)
    with dst.open("wb") as f:
        shutil.copyfileobj(file.file, f)


def coerce_doc_type(value: Optional[str]):
    """Essaie de convertir la chaîne reçue en Enum TypeDocument si possible, sinon renvoie la valeur brute."""
    if not TypeDocument:
        return value  # pas d'enum -> texte libre
    if value is None:
        return getattr(TypeDocument, "AUTRE", list(TypeDocument)[0])
    try:
        return TypeDocument[value]  # par name
    except Exception:
        try:
            return TypeDocument(value)  # par value
        except Exception:
            return getattr(TypeDocument, "AUTRE", list(TypeDocument)[0])

# --------- FORMULAIRE PUBLIC (pour les candidats) ---------
@router.get("/public-form", name="preinscriptions_public_form", response_class=HTMLResponse)
def preinscription_public_form(
    request: Request,
    session: Session = Depends(get_shared_session),
    programme: Optional[str] = Query(None),
):
    try:
        # Récupérer le programme spécifique si fourni
        prog = None
        if programme:
            prog = session.exec(select(Programme).where(Programme.code == programme)).first()
        
        # Récupérer tous les programmes actifs pour la liste déroulante
        programmes_actifs = session.exec(select(Programme).where(Programme.actif.is_(True)).order_by(Programme.code)).all()
        
        return templates.TemplateResponse(
            "pages/programme/preinscription_public_form.html",
            {
                "request": request,
                "settings": settings,
                "programme": prog,
                "programmes_actifs": programmes_actifs,
                "doc_types": DOC_TYPES_DEFAULT,
            },
        )
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        logging.error(f"❌ [ERROR] Erreur dans preinscription_public_form: {e}")
        logging.error(f"❌ [ERROR] Traceback complet:\n{error_trace}")
        
        # Retourner une page d'erreur avec les informations disponibles
        try:
            programmes_actifs = session.exec(select(Programme).where(Programme.actif.is_(True)).order_by(Programme.code)).all()
        except:
            programmes_actifs = []
        
        return templates.TemplateResponse(
            "pages/programme/preinscription_public_form.html",
            {
                "request": request,
                "settings": settings,
                "programme": None,
                "programmes_actifs": programmes_actifs,
                "doc_types": DOC_TYPES_DEFAULT,
                "error": f"Une erreur est survenue lors du chargement du formulaire: {str(e)}"
            },
            status_code=500
        )

# --------- LISTE ADMIN (pour les administrateurs) ---------
@router.get("/form", name="preinscriptions_form", response_class=HTMLResponse)
async def preinscriptions(
    request: Request,
    session: Session = Depends(get_shared_session),
    programme: Optional[str] = Query(None),
    current_user=Depends(get_current_user),
    q: Optional[str] = Query(None),
    schema_routing_service = Depends(get_schema_routing_service),
):
    try:
        # Récupérer le programme depuis request.state (injecté par le middleware)
        programme_code = programme or getattr(request.state, 'current_programme', None) or "ACD"
        
        # Programmes pour filtre
        progs = session.exec(select(Programme).where(Programme.actif.is_(True))).all()
        
        # Récupérer l'objet Programme pour le template
        prog_obj = session.exec(select(Programme).where(Programme.code == programme_code, Programme.actif == True)).first()
        if not prog_obj:
            # Si le programme n'existe pas, utiliser le premier programme actif ou ACD par défaut
            prog_obj = session.exec(select(Programme).where(Programme.code == "ACD", Programme.actif == True)).first()
            if not prog_obj and progs:
                prog_obj = progs[0]
        
        try:
            schema_routing_service.set_schema(programme_code.lower())
        except Exception as schema_error:
            logging.warning(f"Erreur lors de la configuration du schéma {programme_code.lower()}: {schema_error}")
            # Continuer avec le schéma par défaut
        
        # Vérifier l'existence des tables avant d'exécuter les requêtes
        if not table_exists_anywhere("preinscription", session):
            return templates.TemplateResponse("pages/programme/preinscriptions_list.html", {
                "request": request,
                "utilisateur": current_user,
                "programmes": progs,
                "progs": progs,  # Alias pour compatibilité
                "programme": prog_obj,  # Objet Programme pour le template
                "preinscriptions": [],
                "rows": [],  # Alias pour compatibilité
                "total": 0,
                "total_programme": 0,
                "programme_selectionne": programme or programme_code,
                "current_programme": programme or programme_code,
                "q": q or "",
                "settings": settings,
                "kpi": {"total": 0, "programme": 0, "programme_code": programme_code},
                "pins": []
            })
        
        # Construire la requête SQL pour récupérer les préinscriptions avec les jointures
        # Utiliser le schéma du programme pour les tables du programme
        schema = schema_routing_service.get_schema() or "public"
        
        # S'assurer que le search_path est correctement configuré AVANT la requête
        try:
            session.execute(text(f"SET search_path TO {schema}, public"))
        except Exception as e:
            logging.warning(f"⚠️ Erreur lors de la configuration du search_path: {e}")
        
        sql_query = f"""
            SELECT 
                p.id as p_id, p.programme_id, p.candidat_id, p.source, p.statut, p.cree_le,
                p.civilite as p_civilite, p.nom as p_nom, p.prenom as p_prenom, p.date_naissance as p_date_naissance,
                p.email as p_email, p.telephone as p_telephone,
                c.id as c_id, c.civilite, c.nom, c.prenom, c.date_naissance, c.email, c.telephone, c.adresse_personnelle,
                c.lat as c_lat, c.lng as c_lng,
                prog.id as prog_id, prog.code, prog.nom as prog_nom, prog.ca_seuil_min, prog.ca_seuil_max, prog.anciennete_min_annees,
                e.id as e_id, e.adresse, e.lat as e_lat, e.lng as e_lng, e.qpv, e.date_creation as e_date_creation, e.chiffre_affaires as e_chiffre_affaires,
                elig.id as elig_id, elig.verdict, elig.ca_seuil_ok, elig.ca_score, elig.qpv_ok, elig.anciennete_ok, elig.anciennete_annees, elig.calcule_le
            FROM {schema}.preinscription p
            JOIN {schema}.candidat c ON c.id = p.candidat_id
            JOIN public.programme prog ON prog.id = p.programme_id
            LEFT JOIN {schema}.entreprise e ON e.candidat_id = c.id
            LEFT JOIN {schema}.eligibilite elig ON elig.preinscription_id = p.id
        """
        
        params = {}
        conditions = []
        
        if programme:
            conditions.append("prog.code = :programme_code")
            params["programme_code"] = programme
        
        if q:
            conditions.append("(c.nom ILIKE :q OR c.prenom ILIKE :q OR c.email ILIKE :q)")
            params["q"] = f"%{q}%"
        
        if conditions:
            sql_query += " WHERE " + " AND ".join(conditions)
        
        sql_query += " ORDER BY p.cree_le DESC LIMIT 300"
        
        # Exécuter la requête
        try:
            result = session.execute(text(sql_query), params)
            rows = result.fetchall()
            # Convertir les Row en dictionnaires pour faciliter l'accès dans le template
            rows = [dict(row._mapping) for row in rows]
            
            # Si aucune éligibilité n'est récupérée par le LEFT JOIN, vérifier directement dans la base
            elig_count = sum(1 for row in rows if row.get('elig_id') is not None)
            if elig_count == 0 and len(rows) > 0:
                for row in rows[:3]:  # Vérifier les 3 premières
                    p_id = row.get('p_id')
                    try:
                        direct_elig = session.execute(
                            text(f"SELECT id, verdict, ca_seuil_ok, ca_score, qpv_ok, anciennete_ok, anciennete_annees FROM {schema}.eligibilite WHERE preinscription_id = :p_id"),
                            {"p_id": p_id}
                        ).fetchone()
                        if direct_elig:
                            # Mettre à jour la row avec les données trouvées
                            row['elig_id'] = direct_elig[0]
                            row['verdict'] = direct_elig[1]
                            row['ca_seuil_ok'] = direct_elig[2]
                            row['ca_score'] = direct_elig[3]  # Condition CA
                            row['qpv_ok'] = direct_elig[4]
                            row['anciennete_ok'] = direct_elig[5]
                            row['anciennete_annees'] = direct_elig[6]  # Condition ancienneté
                    except Exception as e:
                        logging.warning(f"⚠️ Erreur lors de la vérification directe pour préinscription {p_id}: {e}")
        except Exception as e:
            logging.error(f"Erreur lors de la récupération des préinscriptions: {e}")
            import traceback
            logging.error(traceback.format_exc())
            rows = []

        # Vérifier et calculer l'éligibilité manquante pour chaque préinscription
        from ..services.eligibilite import evaluate_eligibilite, entreprise_age_annees
        from datetime import datetime, timezone
        
        # Vérifier que la table eligibilite existe avant de calculer
        eligibilite_table_exists = table_exists_anywhere('eligibilite', session, schema)
        
        for row in rows:
            preinscription_id = row.get('p_id')
            elig_id = row.get('elig_id')
            
            # Si l'éligibilité n'existe pas, la calculer automatiquement
            if not elig_id and preinscription_id and eligibilite_table_exists:
                try:
                    # Récupérer les données nécessaires depuis la row
                    adresse_personnelle = row.get('adresse_personnelle')
                    adresse_entreprise = row.get('adresse')
                    chiffre_affaires = row.get('e_chiffre_affaires')
                    date_creation_entreprise = row.get('e_date_creation')
                    programme_id = row.get('programme_id')  # ID du programme depuis la row
                    
                    # Calculer l'ancienneté
                    anciennete = entreprise_age_annees(date_creation_entreprise)
                    if anciennete is not None:
                        anciennete = int(anciennete)  # Convertir en entier pour correspondre au type attendu
                    
                    # Convertir le chiffre d'affaires en string si nécessaire
                    ca_string = str(chiffre_affaires) if chiffre_affaires else None
                    
                    # Calculer l'éligibilité
                    verdict, details = await evaluate_eligibilite(
                        adresse_perso=adresse_personnelle,
                        adresse_entreprise=adresse_entreprise,
                        chiffre_affaires=ca_string,
                        anciennete_annees=anciennete,
                        programme_id=programme_id,
                        session=session,
                        request=request,
                        preinscription_id=preinscription_id,
                        schema_name=schema
                    )
                    
                    # L'éligibilité est maintenant enregistrée automatiquement par evaluate_eligibilite
                    # Récupérer les données d'éligibilité pour mettre à jour la row
                    try:
                        session.execute(text(f"SET search_path TO {schema}, public"))
                        elig_result = session.execute(
                            text(f"SELECT id, verdict, ca_seuil_ok, ca_score, qpv_ok, anciennete_ok, anciennete_annees, calcule_le FROM {schema}.eligibilite WHERE preinscription_id = :preinscription_id ORDER BY id DESC LIMIT 1"),
                            {"preinscription_id": preinscription_id}
                        )
                        elig_row = elig_result.fetchone()
                        if elig_row:
                            row['elig_id'] = elig_row[0]
                            row['verdict'] = elig_row[1]
                            row['ca_seuil_ok'] = elig_row[2]
                            row['ca_score'] = elig_row[3]  # Condition CA (ex: "50000 <= 75000 <= 100000")
                            row['qpv_ok'] = elig_row[4]
                            row['anciennete_ok'] = elig_row[5]
                            row['anciennete_annees'] = elig_row[6]  # Condition ancienneté (ex: "2 >= 3")
                            row['calcule_le'] = elig_row[7]
                        else:
                            # Fallback: utiliser les valeurs calculées (convertir les booléens en strings)
                            row['elig_id'] = True
                            row['verdict'] = verdict
                            row['ca_seuil_ok'] = "true" if details.get("ca_ok") else "false"
                            row['ca_score'] = details.get("ca_condition")  # Condition CA depuis details
                            row['qpv_ok'] = details.get("qpv_ok")
                            row['anciennete_ok'] = "true" if details.get("anciennete_ok") else "false"
                            row['anciennete_annees'] = details.get("anciennete_condition")  # Condition ancienneté depuis details
                            row['calcule_le'] = datetime.now(timezone.utc)
                    except Exception as e:
                        logging.warning(f"⚠️ Erreur lors de la récupération de l'éligibilité pour préinscription {preinscription_id}: {e}")
                        # Fallback: utiliser les valeurs calculées (convertir les booléens en strings)
                        row['elig_id'] = True
                        row['verdict'] = verdict
                        row['ca_seuil_ok'] = "true" if details.get("ca_ok") else "false"
                        row['ca_score'] = details.get("ca_condition")  # Condition CA depuis details
                        row['qpv_ok'] = details.get("qpv_ok")
                        row['anciennete_ok'] = "true" if details.get("anciennete_ok") else "false"
                        row['anciennete_annees'] = details.get("anciennete_condition")  # Condition ancienneté depuis details
                        row['calcule_le'] = datetime.now(timezone.utc)
                except Exception as e:
                    logging.warning(f"⚠️ Erreur lors du calcul automatique de l'éligibilité pour préinscription {preinscription_id}: {e}")
                    import traceback
                    logging.warning(traceback.format_exc())
                    # Ne pas faire échouer l'affichage si le calcul échoue
                    session.rollback()
            
        # Vérifier et compléter les données d'éligibilité pour chaque row après traitement
        for row in rows:
            p_id = row.get('p_id')
            elig_id = row.get('elig_id')
            verdict = row.get('verdict')
            
            # Si ni elig_id ni verdict ne sont présents, vérifier s'il existe une éligibilité dans la base
            if not elig_id and not verdict:
                try:
                    session.execute(text(f"SET search_path TO {schema}, public"))
                    elig_result = session.execute(
                        text(f"SELECT id, verdict, ca_seuil_ok, ca_score, qpv_ok, anciennete_ok, anciennete_annees, calcule_le FROM {schema}.eligibilite WHERE preinscription_id = :preinscription_id ORDER BY id DESC LIMIT 1"),
                        {"preinscription_id": p_id}
                    )
                    elig_row = elig_result.fetchone()
                    if elig_row:
                        row['elig_id'] = elig_row[0]
                        row['verdict'] = elig_row[1]
                        row['ca_seuil_ok'] = elig_row[2]
                        row['ca_score'] = elig_row[3]  # Condition CA
                        row['qpv_ok'] = elig_row[4]
                        row['anciennete_ok'] = elig_row[5]
                        row['anciennete_annees'] = elig_row[6]  # Condition ancienneté
                        row['calcule_le'] = elig_row[7]
                except Exception as e:
                    logging.warning(f"⚠️ Erreur lors de la vérification de l'éligibilité pour préinscription {p_id}: {e}")
            
            # Si elig_id existe mais verdict est None, récupérer les données depuis la base
            elif elig_id and not verdict:
                try:
                    session.execute(text(f"SET search_path TO {schema}, public"))
                    elig_result = session.execute(
                        text(f"SELECT id, verdict, ca_seuil_ok, ca_score, qpv_ok, anciennete_ok, anciennete_annees, calcule_le FROM {schema}.eligibilite WHERE id = :elig_id"),
                        {"elig_id": elig_id}
                    )
                    elig_row = elig_result.fetchone()
                    if elig_row:
                        row['verdict'] = elig_row[1]
                        row['ca_seuil_ok'] = elig_row[2]
                        row['ca_score'] = elig_row[3]  # Condition CA
                        row['qpv_ok'] = elig_row[4]
                        row['anciennete_ok'] = elig_row[5]
                        row['anciennete_annees'] = elig_row[6]  # Condition ancienneté
                        row['calcule_le'] = elig_row[7]
                except Exception as e:
                    logging.warning(f"⚠️ Erreur lors de la récupération de l'éligibilité id {elig_id}: {e}")

        # KPI pour en-tête - compter toutes les préinscriptions dans le schéma
        try:
            total_result = session.execute(text(f"SELECT COUNT(*) FROM {schema}.preinscription"))
            total = total_result.fetchone()[0] or 0
        except Exception as e:
            logging.warning(f"Erreur lors du comptage total des préinscriptions: {e}")
            total = 0
        
        # Calculer le total pour le programme sélectionné
        try:
            total_programme_result = session.execute(
                text(f"SELECT COUNT(*) FROM {schema}.preinscription p JOIN public.programme prog ON prog.id = p.programme_id WHERE prog.code = :programme_code"),
                {"programme_code": programme_code}
            )
            total_programme = total_programme_result.fetchone()[0] or 0
        except Exception as e:
            logging.warning(f"Erreur lors du comptage des préinscriptions du programme: {e}")
            total_programme = 0

        # Pins pour carte (lat/lng depuis Entreprise en priorité, sinon depuis Candidat)
        pins = []
        for row in rows:
            try:
                # Accéder aux colonnes par nom (row est maintenant un dictionnaire)
                nom = row.get('nom', '') or ''
                prenom = row.get('prenom', '') or ''
                programme_code_row = row.get('code', '') or ''
                # Priorité : coordonnées de l'entreprise, sinon candidat
                lat = row.get('e_lat') if row.get('e_lat') is not None else row.get('c_lat')
                lng = row.get('e_lng') if row.get('e_lng') is not None else row.get('c_lng')
                qpv = row.get('qpv', False)
                # Priorité : adresse entreprise, sinon adresse personnelle
                adresse = row.get('adresse', '') or row.get('adresse_personnelle', '') or ''
                
                # Vérifier que lat et lng sont des nombres valides
                if lat is not None and lng is not None:
                    try:
                        lat_float = float(lat)
                        lng_float = float(lng)
                        pins.append({
                            "nom": str(nom),
                            "prenom": str(prenom),
                            "programme": str(programme_code_row),
                            "lat": lat_float,
                            "lng": lng_float,
                            "qpv": bool(qpv),
                            "adresse": str(adresse),
                        })
                    except (ValueError, TypeError) as e:
                        logging.warning(f"Valeurs lat/lng invalides: lat={lat}, lng={lng}, erreur: {e}")
                        continue
            except Exception as e:
                logging.warning(f"Erreur lors de la création d'un pin: {e}")
                continue

        # Debug: vérifier le contenu de pins avant sérialisation
        logging.info(f"Pins créés: {len(pins)} éléments")
        for i, pin in enumerate(pins):
            logging.info(f"Pin {i}: {pin}")

        return templates.TemplateResponse(
            "pages/programme/preinscriptions_list.html",
            {
                "request": request,
                "settings": settings,
                "utilisateur": current_user,
                "rows": rows,
                "progs": progs,
                "programmes": progs,  # Alias pour compatibilité
                "programme": prog_obj,  # Objet Programme pour le template
                "current_programme": programme or programme_code,
                "programme_selectionne": programme or programme_code,
                "q": q or "",
                "kpi": {"total": int(total), "programme": int(total_programme), "programme_code": programme_code},
                "pins": pins,
            },
        )
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        logging.error(f"❌ [ERROR] Erreur dans preinscriptions_form: {e}")
        logging.error(f"❌ [ERROR] Traceback complet:\n{error_trace}")
        
        # Retourner une page d'erreur avec les informations disponibles
        try:
            progs = session.exec(select(Programme).where(Programme.actif.is_(True))).all()
            # Récupérer le programme depuis request.state (injecté par le middleware)
            programme_code_error = programme or getattr(request.state, 'current_programme', None) or "ACD"
            # Récupérer l'objet Programme pour le template même en cas d'erreur
            prog_obj_error = session.exec(select(Programme).where(Programme.code == programme_code_error, Programme.actif == True)).first()
            if not prog_obj_error and progs:
                prog_obj_error = progs[0]
        except:
            progs = []
            prog_obj_error = None
            programme_code_error = programme or getattr(request.state, 'current_programme', None) or "ACD"
        
        return templates.TemplateResponse(
            "pages/programme/preinscriptions_list.html",
            {
                "request": request,
                "utilisateur": current_user,
                "settings": settings,
                "programmes": progs,
                "progs": progs,
                "programme": prog_obj_error,  # Objet Programme pour le template
                "rows": [],
                "preinscriptions": [],
                "current_programme": programme_code_error,
                "programme_selectionne": programme_code_error,
                "q": q or "",
                "kpi": {"total": 0, "programme": 0, "programme_code": programme_code_error},
                "pins": [],
                "error_message": f"Une erreur est survenue lors du chargement des préinscriptions: {str(e)}"
            },
            status_code=500
        )

# --------- SOUMISSION PUBLIQUE (sans token) AVEC UPLOAD PHOTO + DOCS ---------
@router.post("/submit", name="preinscription_public_submit")
async def preinscription_public_submit(
    request: Request,
    programme_code: str = Form(...),
    schema_routing_service = Depends(get_schema_routing_service),
    civilite: str= Form(None),
    nom: str = Form(...),
    prenom: str = Form(...),
    date_naissance: str = Form(...),
    email: str = Form(...),
    telephone: str= Form(None),
    # Champs d'adresse personnelle décomposés
    numero_personnel: str= Form(None),
    rue_personnel: str = Form(None),
    code_postal_personnel: str= Form(None),
    ville_personnel: str = Form(None),
    adresse_personnelle: Optional[str] = Form(None),  # Champ consolidé
    
    # Champs d'adresse entreprise décomposés
    numero_entreprise: Optional[str] = Form(None),
    rue_entreprise: Optional[str] = Form(None),
    code_postal_entreprise: Optional[str] = Form(None),
    ville_entreprise: Optional[str] = Form(None),
    adresse_entreprise: Optional[str] = Form(None),  # Champ consolidé
    date_creation_entreprise: Optional[str] = Form(None),
    chiffre_affaire: Optional[str] = Form(None),
    siret: Optional[str] = Form(None),
    niveau_etudes: Optional[str] = Form(None),
    secteur_activite: Optional[str] = Form(None),
    photo_profil: UploadFile | None = File(None),
    session: Session = Depends(get_shared_session),
):
    # Logs de surveillance si debug activé
    if settings.DEBUG:
        print(f"🔍 [DEBUG] Route /preinscriptions/submit appelée")
        print(f"📝 [DEBUG] Données reçues:")
        print(f"   - programme_code: {programme_code}")
        print(f"   - nom: {nom}")
        print(f"   - prenom: {prenom}")
        print(f"   - email: {email}")
        print(f"   - telephone: {telephone}")
        print(f"   - adresse_personnelle: {adresse_personnelle}")
        print(f"   - adresse_entreprise: {adresse_entreprise}")
        print(f"   - photo_profil: {photo_profil.filename if photo_profil else 'Aucune'}")
    
    prog = session.exec(select(Programme).where(Programme.code == programme_code)).first()
    if not prog:
        if settings.DEBUG:
            print(f"❌ [DEBUG] Programme '{programme_code}' introuvable")
        
        # Récupérer tous les programmes actifs pour la liste déroulante
        programmes_actifs = session.exec(select(Programme).where(Programme.actif.is_(True)).order_by(Programme.code)).all()
        
        # Retourner une page d'erreur claire au lieu d'une exception
        return templates.TemplateResponse(
            "pages/programme/preinscription_public_form.html",
            {
                "request": request,
                "settings": settings,
                "programme": None,
                "error": f"Le programme '{programme_code}' n'existe pas dans notre base de données. Veuillez contacter l'administrateur ou choisir un autre programme.",
                "doc_types": DOC_TYPES_DEFAULT,
                "programmes_actifs": programmes_actifs,  # Programmes disponibles
            },
            status_code=400
        )
    
    if settings.DEBUG:
        print(f"✅ [DEBUG] Programme trouvé: {prog.code} - {prog.nom}")

    # Le middleware ProgramSchemaMiddleware crée déjà le schéma automatiquement s'il n'existe pas
    # On vérifie juste que les tables essentielles existent (robustesse)
    schema_name = programme_code.lower()
    schema_service = ProgramSchemaService(session)
    
    # Vérifier si les tables essentielles existent
    required_tables = ['candidat', 'entreprise', 'preinscription', 'eligibilite']
    missing_tables = [t for t in required_tables if not table_exists_anywhere(t, session, schema_name)]
    
    if missing_tables:
        logging.warning(f"Tables manquantes dans le schéma {schema_name}: {missing_tables}, création...")
        schema_service.create_program_schema(programme_code)  # Crée le schéma ET les tables
        session.commit()

    # Le schéma est déjà configuré par le middleware selon l'URL
    # Ne pas le reconfigurer ici pour éviter les conflits
    if settings.DEBUG:
        print(f"🎯 [DEBUG] Utilisation du schéma configuré par le middleware")

    # Validation des adresses avec le schéma Pydantic
    from ..schemas.preinscription_schemas import Adresse
    from pydantic import ValidationError
    
    # Validation adresse personnelle
    try:
        adresse_perso_obj = Adresse(
            address=adresse_personnelle,
            numero=numero_personnel,
            rue=rue_personnel,
            code_postal=code_postal_personnel,
            ville=ville_personnel,
            type_adresse="personnelle"
        )
        adresse_personnelle = adresse_perso_obj.address
        if settings.DEBUG:
            print(f"🏠 [DEBUG] Adresse personnelle validée et formatée: {adresse_personnelle}")
    except ValidationError as e:
        programmes_actifs = session.exec(select(Programme).where(Programme.actif.is_(True)).order_by(Programme.code)).all()
        return templates.TemplateResponse(
            "pages/programme/preinscription_public_form.html",
            {
                "request": request,
                "settings": settings,
                "programme": prog,
                "error": f"Erreur de validation de l'adresse personnelle: {e.errors()[0]['msg']}",
                "doc_types": DOC_TYPES_DEFAULT,
                "programmes_actifs": programmes_actifs,
            },
            status_code=400
        )
    
    # Validation adresse entreprise (optionnelle)
    if any([numero_entreprise, rue_entreprise, code_postal_entreprise, ville_entreprise]):
        try:
            adresse_ent_obj = Adresse(
                address=adresse_entreprise,
                numero=numero_entreprise,
                rue=rue_entreprise,
                code_postal=code_postal_entreprise,
                ville=ville_entreprise,
                type_adresse="entreprise"
            )
            adresse_entreprise = adresse_ent_obj.address
            if settings.DEBUG:
                print(f"🏢 [DEBUG] Adresse entreprise validée et formatée: {adresse_entreprise}")
        except ValidationError as e:
            programmes_actifs = session.exec(select(Programme).where(Programme.actif.is_(True)).order_by(Programme.code)).all()
            return templates.TemplateResponse(
                "programme/preinscription_public_form.html",
                {
                    "request": request,
                    "settings": settings,
                    "programme": prog,
                    "error": f"Erreur de validation de l'adresse entreprise: {e.errors()[0]['msg']}",
                    "doc_types": DOC_TYPES_DEFAULT,
                    "programmes_actifs": programmes_actifs,
                },
                status_code=400
            )

    dn = _date.fromisoformat(date_naissance)
    dce = _date.fromisoformat(date_creation_entreprise) if date_creation_entreprise else None
    # Le chiffre d'affaires est un intervalle (string), pas un nombre
    ca_string = str(chiffre_affaire).strip() if chiffre_affaire else None
    if settings.DEBUG:
        print(f"💰 [DEBUG] Chiffre d'affaires (intervalle): {ca_string}")

    # Obtenir les modèles configurés pour le schéma du programme
    CandidatSchema = schema_routing_service.get_model_for_schema(Candidat, schema_name)
    EntrepriseSchema = schema_routing_service.get_model_for_schema(Entreprise, schema_name)
    PreinscriptionSchema = schema_routing_service.get_model_for_schema(Preinscription, schema_name)
    DocumentSchema = schema_routing_service.get_model_for_schema(Document, schema_name)
    
    # Configurer le schéma dans le service (sans modifier le search_path encore)
    schema_routing_service.current_schema = schema_name
    
    # Configurer le search_path pour le schéma
    try:
        session.execute(text(f"SET search_path TO {schema_name}, public"))
    except Exception as e:
        logging.warning(f"⚠️ Erreur lors de la configuration du search_path: {e}")
        session.rollback()
        session.execute(text(f"SET search_path TO {schema_name}, public"))

    cand = None
    if table_exists_anywhere("candidat", session, schema_name):
        try:
            # Utiliser le modèle avec le schéma configuré
            cand = session.exec(select(CandidatSchema).where(CandidatSchema.email == email)).first()
        except Exception as e:
            logging.warning(f"Erreur lors de la récupération du candidat: {e}")
            session.rollback()
            # Réessayer après rollback
            try:
                session.execute(text(f"SET search_path TO {schema_name}, public"))
                cand = session.exec(select(CandidatSchema).where(CandidatSchema.email == email)).first()
            except Exception as retry_error:
                logging.error(f"Erreur lors de la récupération du candidat après rollback: {retry_error}")
    
    if not cand:
        if settings.DEBUG:
            print(f"🆕 [DEBUG] Création nouveau candidat: {email}")
        try:
            cand = CandidatSchema(email=email, nom=nom, prenom=prenom)
            session.add(cand)
            session.flush()
        except Exception as e:
            logging.error(f"Erreur lors de la création du candidat: {e}")
            session.rollback()
            # Relancer l'exception pour que le gestionnaire d'erreur global la capture
            raise
    else:
        if settings.DEBUG:
            print(f"🔄 [DEBUG] Candidat existant mis à jour: {email}")
    
    # Vérifier si le candidat est déjà inscrit à ce programme (seulement si la table existe)
    existing_inscription = None
    if table_exists_anywhere("inscription", session, schema_name):
        try:
            existing_inscription_result = schema_routing_service.execute_in_schema(
                "SELECT * FROM inscription WHERE candidat_id = :candidat_id AND programme_id = :programme_id",
                {"candidat_id": cand.id, "programme_id": prog.id},
                schema=schema_name
            )
            existing_inscription = existing_inscription_result.fetchone()
        except Exception as e:
            logging.warning(f"Erreur lors de la vérification de l'inscription: {e}")
    
    if existing_inscription:
        if settings.DEBUG:
            print(f"⚠️ [DEBUG] Candidat déjà inscrit au programme {prog.code}")
        
        # Retourner une erreur claire
        programmes_actifs = session.exec(select(Programme).where(Programme.actif.is_(True)).order_by(Programme.code)).all()
        return templates.TemplateResponse(
            "pages/programme/preinscription_public_form.html",
            {
                "request": request,
                "settings": settings,
                "programme": prog,
                "error": f"Vous êtes déjà inscrit au programme '{prog.code} - {prog.nom}'. Vous ne pouvez vous inscrire qu'une seule fois par programme.",
                "doc_types": DOC_TYPES_DEFAULT,
                "programmes_actifs": programmes_actifs,
            },
            status_code=400
        )
    
    # Vérifier si le candidat est déjà préinscrit à ce programme
    existing_preinscription_result = schema_routing_service.execute_in_schema(
        "SELECT * FROM preinscription WHERE candidat_id = :candidat_id AND programme_id = :programme_id",
        {"candidat_id": cand.id, "programme_id": prog.id},
        schema=schema_name
    )
    existing_preinscription = existing_preinscription_result.fetchone()
    
    if existing_preinscription:
        if settings.DEBUG:
            print(f"⚠️ [DEBUG] Candidat déjà préinscrit au programme {prog.code}")
        
        # Retourner une erreur claire
        programmes_actifs = session.exec(select(Programme).where(Programme.actif.is_(True)).order_by(Programme.code)).all()
        return templates.TemplateResponse(
            "pages/programme/preinscription_public_form.html",
            {
                "request": request,
                "settings": settings,
                "programme": prog,
                "error": f"Vous êtes déjà préinscrit au programme '{prog.code} - {prog.nom}'. Vous ne pouvez vous préinscrire qu'une seule fois par programme.",
                "doc_types": DOC_TYPES_DEFAULT,
                "programmes_actifs": programmes_actifs,
            },
            status_code=400
        )
    
    # Mettre à jour le candidat avec toutes les données disponibles (comme dans l'import Excel)
    if civilite:
        cand.civilite = civilite
    if dn:
        cand.date_naissance = dn
    if telephone:
        cand.telephone = telephone
    if adresse_personnelle:
        cand.adresse_personnelle = adresse_personnelle  # Adresse consolidée
    if niveau_etudes:
        cand.niveau_etudes = niveau_etudes
    if secteur_activite:
        cand.secteur_activite = secteur_activite
    
    # Géocoder l'adresse personnelle si elle existe et que les coordonnées ne sont pas déjà présentes
    if adresse_personnelle and (not cand.lat or not cand.lng):
        try:
            latlng = await geocode_one(adresse_personnelle)
            if latlng:
                cand.lat, cand.lng = latlng
                if settings.DEBUG:
                    print(f"✅ [DEBUG] Coordonnées personnelles trouvées: lat={latlng[0]}, lng={latlng[1]}")
        except Exception as e:
            logging.warning(f"Erreur lors du géocodage de l'adresse personnelle: {e}")

    # Créer ou récupérer l'entreprise (comme dans l'import Excel)
    ent = None
    if siret or adresse_entreprise or dce or ca_string:
        # Vérifier si entreprise existe (par SIRET si disponible, sinon par candidat_id)
        if siret:
            ent = session.exec(select(EntrepriseSchema).where(EntrepriseSchema.siret == siret)).first()
        if not ent:
            ent = session.exec(select(EntrepriseSchema).where(EntrepriseSchema.candidat_id == cand.id)).first()
        
        if not ent:
            if settings.DEBUG:
                print(f"🏢 [DEBUG] Création nouvelle entreprise pour candidat {cand.id}")
            ent = EntrepriseSchema(
                candidat_id=cand.id,
                siret=siret,
                adresse=adresse_entreprise,
                date_creation=dce,
                chiffre_affaires=ca_string
            )
            session.add(ent)
            session.flush()
        else:
            if settings.DEBUG:
                print(f"🏢 [DEBUG] Entreprise existante mise à jour pour candidat {cand.id}")
            # Mettre à jour l'entreprise si des données sont fournies
            if adresse_entreprise:
                ent.adresse = adresse_entreprise
            if dce:
                ent.date_creation = dce
            if siret:
                ent.siret = siret
            if ca_string:
                ent.chiffre_affaires = ca_string
        
        # Géocoder l'adresse entreprise si elle existe et que les coordonnées ne sont pas déjà présentes
        if adresse_entreprise and (not ent.lat or not ent.lng):
            try:
                latlng = await geocode_one(adresse_entreprise)
                if latlng:
                    ent.lat, ent.lng = latlng
                    if settings.DEBUG:
                        print(f"✅ [DEBUG] Coordonnées entreprise trouvées: lat={latlng[0]}, lng={latlng[1]}")
            except Exception as e:
                logging.warning(f"Erreur lors du géocodage de l'adresse entreprise: {e}")

    # Obtenir le modèle Preinscription configuré pour le schéma du programme
    PreinscriptionSchema = schema_routing_service.get_model_for_schema(Preinscription, schema_name)
    DocumentSchema = schema_routing_service.get_model_for_schema(Document, schema_name)
    
    # Parser la date de naissance
    date_naissance_parsed = None
    if date_naissance:
        try:
            date_naissance_parsed = datetime.strptime(date_naissance, '%Y-%m-%d').date()
        except:
            pass
    
    # Parser la date de création entreprise
    date_creation_entreprise_parsed = None
    if date_creation_entreprise:
        try:
            date_creation_entreprise_parsed = datetime.strptime(date_creation_entreprise, '%Y-%m-%d').date()
        except:
            pass
    
    # Créer la préinscription dans le schéma du programme en utilisant le modèle SQLModel
    preinscription = PreinscriptionSchema(
        candidat_id=cand.id,
        programme_id=prog.id,
        source="formulaire",
        statut=StatutDossier.SOUMIS,
        cree_le=datetime.now(timezone.utc),
        # Données du candidat
        civilite=civilite,
        nom=nom,
        prenom=prenom,
        date_naissance=date_naissance_parsed,
        email=email,
        telephone=telephone,
        # Adresse personnelle (décomposée)
        numero_personnel=adresse_perso_obj.numero if 'adresse_perso_obj' in locals() else numero_personnel,
        rue_personnel=adresse_perso_obj.rue if 'adresse_perso_obj' in locals() else rue_personnel,
        code_postal_personnel=adresse_perso_obj.code_postal if 'adresse_perso_obj' in locals() else code_postal_personnel,
        ville_personnel=adresse_perso_obj.ville if 'adresse_perso_obj' in locals() else ville_personnel,
        # Adresse entreprise (décomposée)
        numero_entreprise=adresse_ent_obj.numero if 'adresse_ent_obj' in locals() else numero_entreprise,
        rue_entreprise=adresse_ent_obj.rue if 'adresse_ent_obj' in locals() else rue_entreprise,
        code_postal_entreprise=adresse_ent_obj.code_postal if 'adresse_ent_obj' in locals() else code_postal_entreprise,
        ville_entreprise=adresse_ent_obj.ville if 'adresse_ent_obj' in locals() else ville_entreprise,
        # Entreprise
        date_creation_entreprise=date_creation_entreprise_parsed,
        siret=siret,
        chiffre_affaires=ca_string,
        niveau_etudes=niveau_etudes,
        secteur_activite=secteur_activite
    )
    session.add(preinscription)
    session.flush()
    
    # Récupérer l'ID de la préinscription créée
    pre_id = preinscription.id
    
    if settings.DEBUG:
        print(f"📝 [DEBUG] Préinscription créée avec ID: {pre_id} dans le schéma {programme_code.lower()}")

    media_root = ensure_media_root()
    base_dir = media_root / "Preinscrits" / (prog.code or "UNK") / str(pre_id)
    
    if settings.DEBUG:
        print(f"📁 [DEBUG] Dossier média: {base_dir}")

    # Photo (validation + save)
    if photo_profil and getattr(photo_profil, "filename", ""):
        if settings.DEBUG:
            print(f"📸 [DEBUG] Traitement photo: {photo_profil.filename}")
        validate_upload(
            photo_profil,
            allowed_mime_types=settings.ALLOWED_IMAGE_MIME_TYPES,
            max_mb=settings.MAX_UPLOAD_SIZE_MB,
            field_name="photo_profil",
        )
        
        # Utiliser FileUploadService.save_media_file pour sauvegarder dans media/profile_image/{programme_code}/
        file_info = await FileUploadService.save_media_file(
            photo_profil,
            media_type="profile_image",  # Sauvegarde dans media/profile_image/
            programme_code=programme_code,  # Isoler par programme : media/profile_image/{programme_code}/id_{pre_id}/
            subfolder_id=pre_id  # Crée media/profile_image/{programme_code}/id_{pre_id}/
        )
        
        cand.photo_profil = file_info["relative_path"]
        if settings.DEBUG:
            print(f"💾 [DEBUG] Photo sauvegardée: {file_info['relative_path']}")

    # Documents dynamiques
    form = await request.form()
    indices: Set[str] = set()
    for k in form.keys():
        if k.startswith("doc_type_"):
            indices.add(k.split("_")[-1])
    
    if settings.DEBUG:
        print(f"📄 [DEBUG] Documents trouvés: {len(indices)} document(s)")

    for idx in indices:
        doc_type_val = form.get(f"doc_type_{idx}")
        title = form.get(f"doc_title_{idx}")
        file = form.get(f"doc_file_{idx}")  # UploadFile

        if not file or not getattr(file, "filename", ""):
            if settings.DEBUG:
                print(f"⚠️ [DEBUG] Document {idx} ignoré: fichier manquant")
            continue

        if settings.DEBUG:
            print(f"📄 [DEBUG] Traitement document {idx}: {doc_type_val} - {title} - {file.filename}")
        
        validate_upload(
            file,
            allowed_mime_types=settings.ALLOWED_DOC_MIME_TYPES,
            max_mb=settings.MAX_UPLOAD_SIZE_MB,
            field_name=f"doc_file_{idx}",
        )

        doc_type_for_db = coerce_doc_type(doc_type_val)

        doc = DocumentSchema(
            candidat_id=cand.id,
            type_document=doc_type_for_db,
            titre=title,
            nom_fichier=file.filename,
            chemin_fichier="",
            mimetype=getattr(file, "content_type", None),
            taille_octets=None,
            depose_par_id=None,
        )
        session.add(doc)
        session.flush()

        # Utiliser FileUploadService pour sauvegarder le fichier avec isolation par programme
        file_info = await FileUploadService.save_file(
            file,
            "document",  # resource_type
            "Preinscrits",  # folder_name
            programme_code=programme_code,  # Isoler par programme : uploads/Preinscrits/document/{programme_code}/id_{pre_id}/
            subfolder_id=pre_id  # subfolder_id
        )
        
        doc.chemin_fichier = file_info["relative_path"]
        doc.taille_octets = file_info["size_bytes"]

    # Calculer et stocker l'éligibilité (comme dans l'import Excel)
    anciennete = None
    if ent and ent.date_creation:
        anciennete = entreprise_age_annees(ent.date_creation)
        if anciennete is not None:
            anciennete = int(anciennete)  # Convertir en entier pour correspondre au type attendu
    
    # Utiliser le chiffre d'affaires de l'entreprise si disponible, sinon celui de la préinscription
    ca_for_elig = ca_string
    if ent and ent.chiffre_affaires:
        ca_for_elig = str(ent.chiffre_affaires)
    
    # Utiliser l'adresse de l'entreprise si disponible, sinon celle du formulaire
    adresse_entreprise_for_elig = adresse_entreprise
    if ent and ent.adresse:
        adresse_entreprise_for_elig = ent.adresse
    
    verdict, details = await evaluate_eligibilite(
        adresse_perso=adresse_personnelle,
        adresse_entreprise=adresse_entreprise_for_elig,
        chiffre_affaires=ca_for_elig,
        anciennete_annees=anciennete,
        programme_id=prog.id,
        session=session,
        request=request,
        preinscription_id=pre_id,
        schema_name=schema_name
    )
    
    # L'éligibilité est maintenant enregistrée automatiquement par evaluate_eligibilite
    # Plus besoin d'insertion manuelle
    
    # Mettre à jour QPV dans l'entreprise si disponible depuis evaluate_eligibilite
    if ent and details.get("qpv_ok"):
        qpv_nom_from_details = details.get("qpv_ok")
        if qpv_nom_from_details.startswith("QPV"):
            ent.qpv = True
            logging.info(f"✅ QPV trouvé dans evaluate_eligibilite: {qpv_nom_from_details}, mis à jour dans l'entreprise")
    
    # Commit final pour finaliser toutes les modifications
    session.commit()
    
    if settings.DEBUG:
        print(f"✅ [DEBUG] Préinscription terminée avec succès!")
        print(f"🎯 [DEBUG] Redirection vers: /preinscriptions/merci?programme={programme_code}")
    
    # Rediriger vers la page de remerciement avec le programme dans les query params
    redirect_url = request.url_for("preinscriptions_merci")
    redirect_url = f"{redirect_url}?programme={programme_code}"
    return RedirectResponse(url=redirect_url, status_code=303)

# --------- PAGE MERCI ---------
@router.get("/merci", name="preinscriptions_merci", response_class=HTMLResponse)
async def preinscription_merci(request: Request):
    """
    Page de remerciement après soumission du formulaire de préinscription
    """
    try:
        # Récupérer le programme depuis les query params si disponible
        programme_code = request.query_params.get("programme", "ACD")
        
        return templates.TemplateResponse(
            "pages/programme/preinscription_merci.html",
            {
                "request": request,
                "settings": settings,
                "current_programme": programme_code,
                "programme": None,  # Pas besoin du programme complet pour cette page
            },
        )
    except Exception as e:
        logging.error(f"Erreur lors du rendu de la page merci: {e}")
        import traceback
        logging.error(traceback.format_exc())
        # Retourner une page d'erreur simplifiée
        return templates.TemplateResponse(
            "pages/programme/preinscription_merci.html",
            {
                "request": request,
                "settings": settings,
                "current_programme": "ACD",
                "programme": None,
            },
        )

# --------- TÉLÉCHARGER TEMPLATE EXCEL ---------
@router.get("/download-template", name="preinscriptions_download_template")
def download_template(
    request: Request,
    session: Session = Depends(get_shared_session),
    current_user=Depends(get_current_user),
):
    """Génère et télécharge un template Excel pour l'import de préinscriptions"""
    try:
        # Créer un nouveau workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Préinscriptions"
        
        # Définir les en-têtes de colonnes
        headers = [
            "Civilité",
            "Nom",
            "Prénom",
            "Date de naissance (YYYY-MM-DD)",
            "Email",
            "Téléphone",
            "Numéro (adresse personnelle)",
            "Rue (adresse personnelle)",
            "Code postal (adresse personnelle)",
            "Ville (adresse personnelle)",
            "Numéro (adresse entreprise)",
            "Rue (adresse entreprise)",
            "Code postal (adresse entreprise)",
            "Ville (adresse entreprise)",
            "Date de création entreprise (YYYY-MM-DD)",
            "SIRET",
            "Chiffre d'affaires",
            "Niveau d'études",
            "Secteur d'activité",
            "Spécialité culinaire",
            "Nom du concept",
            "Site internet",
            "Réseaux sociaux",
            "Handicap (true/false)",
            "QPV (true/false)"
        ]
        
        # Écrire les en-têtes
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        
        # Ajouter une ligne d'exemple
        example_row = [
            "M.",
            "Dupont",
            "Jean",
            "1990-01-15",
            "jean.dupont@example.com",
            "0612345678",
            "10",
            "Rue de la Paix",
            "75001",
            "Paris",
            "5",
            "Avenue des Champs",
            "75008",
            "Paris",
            "2020-06-15",
            "12345678901234",
            "50000",
            "Bac+5",
            "Informatique",
            "",
            "",
            "",
            "",
            "false",
            "false"
        ]
        
        for col_idx, value in enumerate(example_row, start=1):
            cell = ws.cell(row=2, column=col_idx, value=value)
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        
        # Ajuster la largeur des colonnes (ajout de 4 colonnes supplémentaires)
        column_widths = [12, 20, 20, 25, 30, 15, 12, 25, 18, 20, 12, 25, 18, 20, 30, 18, 18, 18, 25, 25, 25, 25, 25, 18, 18]
        for col_idx, width in enumerate(column_widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        
        # Créer un fichier en mémoire
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Retourner le fichier avec les bons headers
        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": 'attachment; filename="template_preinscriptions.xlsx"',
                "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            }
        )
    except Exception as e:
        logging.error(f"Erreur lors de la génération du template Excel: {e}")
        raise HTTPException(status_code=500, detail=f"Erreur lors de la génération du template: {str(e)}")

# --------- IMPORTER EXCEL ---------
@router.post("/import-excel", name="preinscriptions_import_excel")
async def import_excel(
    request: Request,
    excel_file: UploadFile = File(...),
    programme_code: Optional[str] = Form(None),
    session: Session = Depends(get_shared_session),
    current_user=Depends(get_current_user),
    schema_routing_service = Depends(get_schema_routing_service),
):
    """Importe des préinscriptions depuis un fichier Excel"""
    logging.info(f"📥 Import Excel démarré - programme_code: {programme_code}, fichier: {excel_file.filename}")
    imported_count = 0
    errors = []
    warnings = []
    
    try:
        # Lire le fichier Excel
        try:
            logging.info(f"📖 Lecture du fichier Excel: {excel_file.filename}")
            df = pd.read_excel(excel_file.file, engine='openpyxl')
            logging.info(f"📊 Fichier Excel lu: {len(df)} ligne(s) trouvée(s)")
        except Exception as e:
            error_msg = f"Erreur lors de la lecture du fichier Excel: {str(e)}"
            logging.error(error_msg)
            import traceback
            logging.error(traceback.format_exc())
            raise HTTPException(status_code=400, detail=error_msg)
        
        
        # Vérifier que le fichier n'est pas vide
        if df.empty:
            error_msg = "Le fichier Excel est vide ou ne contient aucune donnée"
            raise HTTPException(status_code=400, detail=error_msg)
        
        # Normaliser les noms de colonnes (enlever espaces en début/fin)
        df.columns = df.columns.str.strip()
        
        # Vérifier que les colonnes essentielles sont présentes
        required_columns = ['Nom', 'Prénom', 'Email']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            error_msg = f"Colonnes manquantes dans le fichier Excel: {', '.join(missing_columns)}"
            raise HTTPException(status_code=400, detail=error_msg)
        
        # Supprimer les lignes vides (ligne d'exemple et lignes sans données essentielles)
        initial_count = len(df)
        df = df.dropna(subset=['Nom', 'Prénom', 'Email'], how='all')
        dropped_count = initial_count - len(df)
        if dropped_count > 0:
            warnings.append(f"{dropped_count} ligne(s) vide(s) ignorée(s)")
        
        
        if len(df) == 0:
            error_msg = "Aucune ligne valide trouvée dans le fichier Excel"
            raise HTTPException(status_code=400, detail=error_msg)
        
        logging.info(f"🔄 Début du traitement de {len(df)} ligne(s)")
        
        # Traiter chaque ligne
        for idx, row in df.iterrows():
            try:
                
                # Utiliser le programme_code du formulaire directement (déjà extrait par FastAPI)
                # ou depuis les query params si non fourni
                prog_code = programme_code or request.query_params.get('programme') or 'ACD'
                prog_code = prog_code.upper() if prog_code != 'PUBLIC' else 'ACD'
                
                # Récupérer et valider le programme
                prog = session.exec(select(Programme).where(Programme.code == prog_code, Programme.actif == True)).first()
                if not prog or prog_code == 'PUBLIC':
                    error_msg = f"Ligne {idx + 2}: Programme '{prog_code}' introuvable ou inactif"
                    logging.error(error_msg)
                    errors.append(error_msg)
                    continue
                
                
                # Le middleware ProgramSchemaMiddleware crée déjà le schéma automatiquement s'il n'existe pas
                # On vérifie juste que les tables essentielles existent (robustesse)
                schema_name = prog_code.lower()
                schema_service = ProgramSchemaService(session)
                
                # Vérifier si les tables essentielles existent
                required_tables = ['candidat', 'entreprise', 'preinscription', 'eligibilite']
                missing_tables = [t for t in required_tables if not table_exists_anywhere(t, session, schema_name)]
                
                if missing_tables:
                    logging.warning(f"Tables manquantes dans le schéma {schema_name}: {missing_tables}, création...")
                    schema_service.create_program_schema(prog_code)  # Crée le schéma ET les tables
                    session.commit()
                else:
                    # Vérifier et ajouter les colonnes manquantes si nécessaire
                    # Comparer les colonnes du modèle SQLModel avec celles de la table existante
                    try:
                        from sqlalchemy import inspect as sa_inspect
                        inspector = sa_inspect(session.bind)
                        
                        # Obtenir les colonnes du modèle Candidat depuis SQLAlchemy
                        if hasattr(Candidat, '__table__') and Candidat.__table__ is not None:
                            model_columns = {col.name: str(col.type) for col in Candidat.__table__.columns}
                            
                            # Vérifier les colonnes de la table candidat existante
                            if inspector.has_table('candidat', schema=schema_name):
                                existing_columns = {col['name']: str(col['type']) for col in inspector.get_columns('candidat', schema=schema_name)}
                                missing_columns = {col: sql_type for col, sql_type in model_columns.items() if col not in existing_columns}
                                
                                if missing_columns:
                                    logging.info(f"Colonnes manquantes dans candidat ({schema_name}): {list(missing_columns.keys())}, ajout...")
                                    for col_name, sql_type in missing_columns.items():
                                        # Obtenir la définition complète de la colonne depuis le modèle
                                        col_def = Candidat.__table__.columns[col_name]
                                        
                                        # Convertir les types ENUM en VARCHAR pour éviter les erreurs
                                        # IMPORTANT: Les enums sont stockés en VARCHAR dans la base, pas comme des types ENUM PostgreSQL
                                        sql_type_str = str(sql_type)
                                        # Vérifier si c'est un type ENUM (peut être représenté comme ENUM ou EnumType dans SQLAlchemy)
                                        if 'ENUM' in sql_type_str.upper() or 'Enum' in sql_type_str or hasattr(col_def.type, 'enums'):
                                            # Si c'est un enum, utiliser VARCHAR(50) par défaut
                                            sql_type_str = "VARCHAR(50)"
                                        
                                        default_clause = ""
                                        
                                        # Générer la clause DEFAULT si nécessaire
                                        if col_def.default is not None:
                                            default_value = col_def.default.arg if hasattr(col_def.default, 'arg') else col_def.default
                                            if isinstance(default_value, bool):
                                                default_clause = f" DEFAULT {str(default_value).upper()}"
                                            elif isinstance(default_value, (int, float)):
                                                default_clause = f" DEFAULT {default_value}"
                                            elif isinstance(default_value, str):
                                                default_clause = f" DEFAULT '{default_value}'"
                                        elif col_def.nullable:
                                            default_clause = " DEFAULT NULL"
                                        
                                        session.execute(text(f"ALTER TABLE {schema_name}.candidat ADD COLUMN IF NOT EXISTS {col_name} {sql_type_str}{default_clause}"))
                                    session.commit()
                            
                            # Vérifier et ajouter les colonnes manquantes dans la table preinscription
                            if hasattr(Preinscription, '__table__') and Preinscription.__table__ is not None:
                                model_columns_preinscription = {col.name: str(col.type) for col in Preinscription.__table__.columns}
                                
                                # Vérifier les colonnes de la table preinscription existante
                                if inspector.has_table('preinscription', schema=schema_name):
                                    existing_columns_preinscription = {col['name']: str(col['type']) for col in inspector.get_columns('preinscription', schema=schema_name)}
                                    
                                    # Supprimer la colonne donnees_brutes_json si elle existe encore (elle ne devrait plus exister)
                                    if 'donnees_brutes_json' in existing_columns_preinscription:
                                        try:
                                            session.execute(text(f"ALTER TABLE {schema_name}.preinscription DROP COLUMN IF EXISTS donnees_brutes_json"))
                                            session.commit()
                                            # Mettre à jour la liste des colonnes existantes
                                            existing_columns_preinscription = {col['name']: str(col['type']) for col in inspector.get_columns('preinscription', schema=schema_name)}
                                        except Exception as e:
                                            logging.warning(f"Erreur lors de la suppression de 'donnees_brutes_json': {e}")
                                    
                                    missing_columns_preinscription = {col: sql_type for col, sql_type in model_columns_preinscription.items() if col not in existing_columns_preinscription}
                                    
                                    if missing_columns_preinscription:
                                        logging.info(f"Colonnes manquantes dans preinscription ({schema_name}): {list(missing_columns_preinscription.keys())}, ajout...")
                                        for col_name, sql_type in missing_columns_preinscription.items():
                                            # Obtenir la définition complète de la colonne depuis le modèle
                                            col_def = Preinscription.__table__.columns[col_name]
                                            
                                            # Convertir les types ENUM en VARCHAR pour éviter les erreurs
                                            sql_type_str = str(sql_type)
                                            if 'ENUM' in sql_type_str.upper() or 'Enum' in sql_type_str or hasattr(col_def.type, 'enums'):
                                                sql_type_str = "VARCHAR(50)"
                                            # Pour les dates, utiliser DATE
                                            elif 'DATE' in sql_type_str.upper():
                                                sql_type_str = "DATE"
                                            # Pour les timestamps, utiliser TIMESTAMP WITH TIME ZONE
                                            elif 'TIMESTAMP' in sql_type_str.upper() or 'DATETIME' in sql_type_str.upper():
                                                sql_type_str = "TIMESTAMP WITH TIME ZONE"
                                            # Pour les chaînes de caractères longues, utiliser TEXT ou VARCHAR selon le cas
                                            elif 'TEXT' in sql_type_str.upper():
                                                sql_type_str = "TEXT"
                                            # Pour les autres VARCHAR, garder le type tel quel
                                            elif 'VARCHAR' in sql_type_str.upper():
                                                # Garder le VARCHAR avec sa longueur si spécifiée
                                                sql_type_str = sql_type_str
                                            else:
                                                # Pour les autres types, utiliser le type tel quel
                                                sql_type_str = sql_type_str
                                            
                                            default_clause = ""
                                            
                                            # Générer la clause DEFAULT si nécessaire
                                            if col_def.default is not None:
                                                default_value = col_def.default.arg if hasattr(col_def.default, 'arg') else col_def.default
                                                # Gérer les valeurs par défaut pour les enums (valeurs de l'enum)
                                                if hasattr(default_value, 'value'):
                                                    default_value = default_value.value
                                                if isinstance(default_value, bool):
                                                    default_clause = f" DEFAULT {str(default_value).upper()}"
                                                elif isinstance(default_value, (int, float)):
                                                    default_clause = f" DEFAULT {default_value}"
                                                elif isinstance(default_value, str):
                                                    default_clause = f" DEFAULT '{default_value}'"
                                            elif col_def.nullable:
                                                default_clause = " DEFAULT NULL"
                                            
                                            session.execute(text(f"ALTER TABLE {schema_name}.preinscription ADD COLUMN IF NOT EXISTS {col_name} {sql_type_str}{default_clause}"))
                                        session.commit()
                    except Exception as e:
                        logging.warning(f"Erreur lors de la vérification des colonnes: {e}")
                
                # Configurer le schéma pour cette session avec SchemaRoutingService
                # IMPORTANT: Ne pas appeler set_schema() ici car cela configure le search_path
                # qui peut faire que SQLAlchemy ignore le schéma explicite de la table
                # À la place, on configure juste le schéma dans le service sans modifier le search_path
                schema_routing_service.current_schema = schema_name
                # schema_routing_service.set_schema(schema_name)  # NE PAS APPELER car cela configure le search_path
                
                # Utiliser get_model_for_schema pour obtenir les modèles configurés pour le schéma
                CandidatSchema = schema_routing_service.get_model_for_schema(Candidat, schema_name)
                EntrepriseSchema = schema_routing_service.get_model_for_schema(Entreprise, schema_name)
                PreinscriptionSchema = schema_routing_service.get_model_for_schema(Preinscription, schema_name)
                
                # IMPORTANT: Configurer le search_path APRÈS avoir obtenu les modèles
                # pour que SQLAlchemy utilise le schéma explicite de la table
                session.execute(text(f"SET search_path TO {schema_name}, public"))
                
                # Données du candidat
                email = str(row.get('Email', '')).strip()
                if not email or email == 'nan':
                    error_msg = f"Ligne {idx + 2}: Email manquant"
                    errors.append(error_msg)
                    continue
                
                nom = str(row.get('Nom', '')).strip()
                prenom = str(row.get('Prénom', '')).strip()
                
                if not nom or nom == 'nan' or not prenom or prenom == 'nan':
                    error_msg = f"Ligne {idx + 2}: Nom ou prénom manquant"
                    errors.append(error_msg)
                    continue
                
                # Extraire toutes les données du candidat depuis le fichier Excel
                civilite = None
                if pd.notna(row.get('Civilité')):
                    civilite = str(row.get('Civilité')).strip()
                    if civilite == 'nan':
                        civilite = None
                
                date_naissance = None
                date_str = row.get('Date de naissance (YYYY-MM-DD)')
                if pd.notna(date_str):
                    if isinstance(date_str, str):
                        try:
                            date_naissance = datetime.strptime(date_str.strip(), '%Y-%m-%d').date()
                        except:
                            pass
                    elif hasattr(date_str, 'date'):
                        date_naissance = date_str.date()
                
                telephone = None
                if pd.notna(row.get('Téléphone')):
                    telephone = str(row.get('Téléphone')).strip()
                    if telephone == 'nan':
                        telephone = None
                
                # Construire l'adresse personnelle à partir des composants
                numero_personnel = None
                rue_personnel = None
                code_postal_personnel = None
                ville_personnel = None
                
                if pd.notna(row.get('Numéro (adresse personnelle)')):
                    numero_personnel = str(row.get('Numéro (adresse personnelle)')).strip()
                if pd.notna(row.get('Rue (adresse personnelle)')):
                    rue_personnel = str(row.get('Rue (adresse personnelle)')).strip()
                if pd.notna(row.get('Code postal (adresse personnelle)')):
                    code_postal_personnel = str(row.get('Code postal (adresse personnelle)')).strip()
                if pd.notna(row.get('Ville (adresse personnelle)')):
                    ville_personnel = str(row.get('Ville (adresse personnelle)')).strip()
                
                adresse_personnelle_parts = [p for p in [numero_personnel, rue_personnel, code_postal_personnel, ville_personnel] if p and p != 'nan']
                adresse_personnelle = ', '.join(adresse_personnelle_parts) if adresse_personnelle_parts else None
                
                niveau_etudes = None
                if pd.notna(row.get("Niveau d'études")):
                    niveau_etudes = str(row.get("Niveau d'études")).strip()
                    if niveau_etudes == 'nan':
                        niveau_etudes = None
                
                secteur_activite = None
                if pd.notna(row.get("Secteur d'activité")):
                    secteur_activite = str(row.get("Secteur d'activité")).strip()
                    if secteur_activite == 'nan':
                        secteur_activite = None
                
                # Récupérer ou créer le candidat
                # Vérifier que le modèle a bien le schéma configuré
                if hasattr(CandidatSchema, '__table__'):
                    table_schema = getattr(CandidatSchema.__table__, 'schema', None)
                    
                    # IMPORTANT: Forcer SQLAlchemy à utiliser le schéma en vérifiant que la table est bien configurée
                    # Si le schéma n'est pas dans le SQL généré, c'est que SQLAlchemy ignore le schéma de la table
                    # On doit s'assurer que la table a bien le schéma défini ET que SQLAlchemy l'utilise
                    if table_schema != schema_name:
                        CandidatSchema.__table__.schema = schema_name
                
                # IMPORTANT: Réappliquer le search_path juste avant la requête pour s'assurer qu'il est bien configuré
                session.execute(text(f"SET search_path TO {schema_name}, public"))
                
                cand = session.exec(select(CandidatSchema).where(CandidatSchema.email == email)).first()
                if not cand:
                    cand = CandidatSchema(email=email, nom=nom, prenom=prenom)
                    session.add(cand)
                    session.flush()
                
                # Vérifier si une préinscription existe déjà pour ce candidat et ce programme
                existing_pre = session.exec(
                    select(PreinscriptionSchema)
                    .join(CandidatSchema)
                    .where(CandidatSchema.email == email, PreinscriptionSchema.programme_id == prog.id)
                ).first()
                
                if existing_pre:
                    error_msg = f"Ligne {idx + 2}: Préinscription déjà existante pour {email} au programme {prog_code}"
                    errors.append(error_msg)
                    continue
                
                # Extraire les données d'adresse personnelle
                numero_personnel = None
                rue_personnel = None
                code_postal_personnel = None
                ville_personnel = None
                
                if pd.notna(row.get('Numéro (adresse personnelle)')):
                    numero_personnel = str(row.get('Numéro (adresse personnelle)')).strip()
                    if numero_personnel == 'nan':
                        numero_personnel = None
                if pd.notna(row.get('Rue (adresse personnelle)')):
                    rue_personnel = str(row.get('Rue (adresse personnelle)')).strip()
                    if rue_personnel == 'nan':
                        rue_personnel = None
                if pd.notna(row.get('Code postal (adresse personnelle)')):
                    code_postal_personnel = str(row.get('Code postal (adresse personnelle)')).strip()
                    if code_postal_personnel == 'nan':
                        code_postal_personnel = None
                if pd.notna(row.get('Ville (adresse personnelle)')):
                    ville_personnel = str(row.get('Ville (adresse personnelle)')).strip()
                    if ville_personnel == 'nan':
                        ville_personnel = None
                
                # Extraire les données d'adresse entreprise
                numero_entreprise_preinscription = None
                rue_entreprise_preinscription = None
                code_postal_entreprise_preinscription = None
                ville_entreprise_preinscription = None
                
                if pd.notna(row.get('Numéro (adresse entreprise)')):
                    numero_entreprise_preinscription = str(row.get('Numéro (adresse entreprise)')).strip()
                    if numero_entreprise_preinscription == 'nan':
                        numero_entreprise_preinscription = None
                if pd.notna(row.get('Rue (adresse entreprise)')):
                    rue_entreprise_preinscription = str(row.get('Rue (adresse entreprise)')).strip()
                    if rue_entreprise_preinscription == 'nan':
                        rue_entreprise_preinscription = None
                if pd.notna(row.get('Code postal (adresse entreprise)')):
                    code_postal_entreprise_preinscription = str(row.get('Code postal (adresse entreprise)')).strip()
                    if code_postal_entreprise_preinscription == 'nan':
                        code_postal_entreprise_preinscription = None
                if pd.notna(row.get('Ville (adresse entreprise)')):
                    ville_entreprise_preinscription = str(row.get('Ville (adresse entreprise)')).strip()
                    if ville_entreprise_preinscription == 'nan':
                        ville_entreprise_preinscription = None
                
                # Extraire date_creation_entreprise pour la préinscription
                date_creation_entreprise_preinscription = None
                date_str_entreprise = row.get('Date de création entreprise (YYYY-MM-DD)')
                if pd.notna(date_str_entreprise):
                    if isinstance(date_str_entreprise, str):
                        try:
                            date_creation_entreprise_preinscription = datetime.strptime(date_str_entreprise.strip(), '%Y-%m-%d').date()
                        except:
                            pass
                    elif hasattr(date_str_entreprise, 'date'):
                        date_creation_entreprise_preinscription = date_str_entreprise.date()
                
                # Extraire siret pour la préinscription
                siret_preinscription = None
                if pd.notna(row.get('SIRET')):
                    siret_preinscription = str(row.get('SIRET')).strip()
                    if siret_preinscription == 'nan':
                        siret_preinscription = None
                
                # Extraire chiffre_affaires pour la préinscription
                chiffre_affaires_preinscription = None
                if pd.notna(row.get("Chiffre d'affaires")):
                    chiffre_affaires_preinscription = str(row.get("Chiffre d'affaires")).strip()
                    if chiffre_affaires_preinscription == 'nan':
                        chiffre_affaires_preinscription = None
                
                # Extraire les champs de restauration
                specialite_culinaire = None
                if pd.notna(row.get('Spécialité culinaire')):
                    specialite_culinaire = str(row.get('Spécialité culinaire')).strip()
                    if specialite_culinaire == 'nan':
                        specialite_culinaire = None
                
                nom_concept = None
                if pd.notna(row.get('Nom du concept')):
                    nom_concept = str(row.get('Nom du concept')).strip()
                    if nom_concept == 'nan':
                        nom_concept = None
                
                site_internet = None
                if pd.notna(row.get('Site internet')):
                    site_internet = str(row.get('Site internet')).strip()
                    if site_internet == 'nan':
                        site_internet = None
                
                lien_reseaux_sociaux = None
                if pd.notna(row.get('Réseaux sociaux')):
                    lien_reseaux_sociaux = str(row.get('Réseaux sociaux')).strip()
                    if lien_reseaux_sociaux == 'nan':
                        lien_reseaux_sociaux = None
                
                # Extraire les champs handicap et QPV
                handicap = False
                if pd.notna(row.get('Handicap (true/false)')):
                    handicap_str = str(row.get('Handicap (true/false)')).strip().lower()
                    handicap = handicap_str in ['true', '1', 'oui', 'yes', 'vrai']
                
                qpv = False
                if pd.notna(row.get('QPV (true/false)')):
                    qpv_str = str(row.get('QPV (true/false)')).strip().lower()
                    qpv = qpv_str in ['true', '1', 'oui', 'yes', 'vrai']
                
                
                
                preinscription = PreinscriptionSchema(
                    candidat_id=cand.id,
                    programme_id=prog.id,
                    source="import",  # Indiquer que c'est un import Excel
                    statut=StatutDossier.SOUMIS,  # Statut par défaut pour les imports
                    cree_le=datetime.now(timezone.utc),
                    # Données du candidat
                    civilite=civilite,
                    nom=nom,
                    prenom=prenom,
                    date_naissance=date_naissance,
                    email=email,
                    telephone=telephone,
                    # Adresse personnelle (décomposée)
                    numero_personnel=numero_personnel,
                    rue_personnel=rue_personnel,
                    code_postal_personnel=code_postal_personnel,
                    ville_personnel=ville_personnel,
                    # Adresse entreprise (décomposée)
                    numero_entreprise=numero_entreprise_preinscription,
                    rue_entreprise=rue_entreprise_preinscription,
                    code_postal_entreprise=code_postal_entreprise_preinscription,
                    ville_entreprise=ville_entreprise_preinscription,
                    # Entreprise
                    date_creation_entreprise=date_creation_entreprise_preinscription,
                    siret=siret_preinscription,
                    chiffre_affaires=chiffre_affaires_preinscription,
                    niveau_etudes=niveau_etudes,
                    secteur_activite=secteur_activite
                )
                session.add(preinscription)
                session.flush()
                
                # Mettre à jour le candidat avec toutes les données disponibles (toujours dans le schéma du programme)
                if civilite:
                    cand.civilite = civilite
                if date_naissance:
                    cand.date_naissance = date_naissance
                if telephone:
                    cand.telephone = telephone
                if adresse_personnelle:
                    cand.adresse_personnelle = adresse_personnelle
                if niveau_etudes:
                    cand.niveau_etudes = niveau_etudes
                if secteur_activite:
                    cand.secteur_activite = secteur_activite
                # Assigner handicap au candidat
                cand.handicap = handicap
                
                # Géocoder l'adresse personnelle si elle existe et que les coordonnées ne sont pas déjà présentes
                if adresse_personnelle and (not cand.lat or not cand.lng):
                    try:
                        latlng = await geocode_one(adresse_personnelle)
                        if latlng:
                            cand.lat, cand.lng = latlng
                    except Exception as e:
                        logging.warning(f"Erreur lors du géocodage de l'adresse personnelle (ligne {idx + 2}): {e}")
                
                # Créer l'entreprise si des données sont fournies (toujours dans le schéma du programme)
                siret = None
                if pd.notna(row.get('SIRET')):
                    siret = str(row.get('SIRET')).strip()
                    if siret == 'nan':
                        siret = None
                
                # Vérifier si on doit créer/mettre à jour l'entreprise (SIRET ou données de restauration présentes)
                doit_creer_entreprise = (siret and siret != 'nan') or specialite_culinaire or nom_concept or site_internet or lien_reseaux_sociaux
                
                if doit_creer_entreprise:
                    # Vérifier si entreprise existe (par SIRET si disponible, sinon par candidat_id)
                    entreprise = None
                    if siret and siret != 'nan':
                        entreprise = session.exec(select(EntrepriseSchema).where(EntrepriseSchema.siret == siret)).first()
                    if not entreprise:
                        entreprise = session.exec(select(EntrepriseSchema).where(EntrepriseSchema.candidat_id == cand.id)).first()
                    if not entreprise:
                        numero_entreprise = None
                        rue_entreprise = None
                        code_postal_entreprise = None
                        ville_entreprise = None
                        
                        if pd.notna(row.get('Numéro (adresse entreprise)')):
                            numero_entreprise = str(row.get('Numéro (adresse entreprise)')).strip()
                        if pd.notna(row.get('Rue (adresse entreprise)')):
                            rue_entreprise = str(row.get('Rue (adresse entreprise)')).strip()
                        if pd.notna(row.get('Code postal (adresse entreprise)')):
                            code_postal_entreprise = str(row.get('Code postal (adresse entreprise)')).strip()
                        if pd.notna(row.get('Ville (adresse entreprise)')):
                            ville_entreprise = str(row.get('Ville (adresse entreprise)')).strip()
                        
                        adresse_entreprise_parts = [p for p in [numero_entreprise, rue_entreprise, code_postal_entreprise, ville_entreprise] if p]
                        adresse_entreprise = ', '.join(adresse_entreprise_parts) if adresse_entreprise_parts else None
                        
                        date_creation_entreprise = None
                        date_str = row.get('Date de création entreprise (YYYY-MM-DD)')
                        if pd.notna(date_str):
                            if isinstance(date_str, str):
                                try:
                                    date_creation_entreprise = datetime.strptime(date_str.strip(), '%Y-%m-%d').date()
                                except:
                                    pass
                            elif hasattr(date_str, 'date'):
                                date_creation_entreprise = date_str.date()
                        
                        # Extraire le chiffre d'affaires depuis le fichier Excel
                        chiffre_affaires = None
                        if pd.notna(row.get("Chiffre d'affaires")):
                            chiffre_affaires = str(row.get("Chiffre d'affaires")).strip()
                            if chiffre_affaires == 'nan':
                                chiffre_affaires = None
                        
                        entreprise = EntrepriseSchema(
                            candidat_id=cand.id,
                            siret=siret,
                            adresse=adresse_entreprise,
                            date_creation=date_creation_entreprise,
                            chiffre_affaires=chiffre_affaires,
                            specialite_culinaire=specialite_culinaire,
                            nom_concept=nom_concept,
                            site_internet=site_internet,
                            lien_reseaux_sociaux=lien_reseaux_sociaux,
                            qpv=qpv
                        )
                        session.add(entreprise)
                        session.flush()  # Flush pour obtenir l'ID de l'entreprise
                        
                        # Géocoder l'adresse entreprise si elle existe et que les coordonnées ne sont pas déjà présentes
                        if adresse_entreprise and (not entreprise.lat or not entreprise.lng):
                            try:
                                latlng = await geocode_one(adresse_entreprise)
                                if latlng:
                                    entreprise.lat, entreprise.lng = latlng
                            except Exception as e:
                                logging.warning(f"Erreur lors du géocodage de l'adresse entreprise (ligne {idx + 2}): {e}")
                    else:
                        # Entreprise existe déjà, mettre à jour si nécessaire
                        numero_entreprise = None
                        rue_entreprise = None
                        code_postal_entreprise = None
                        ville_entreprise = None
                        
                        if pd.notna(row.get('Numéro (adresse entreprise)')):
                            numero_entreprise = str(row.get('Numéro (adresse entreprise)')).strip()
                        if pd.notna(row.get('Rue (adresse entreprise)')):
                            rue_entreprise = str(row.get('Rue (adresse entreprise)')).strip()
                        if pd.notna(row.get('Code postal (adresse entreprise)')):
                            code_postal_entreprise = str(row.get('Code postal (adresse entreprise)')).strip()
                        if pd.notna(row.get('Ville (adresse entreprise)')):
                            ville_entreprise = str(row.get('Ville (adresse entreprise)')).strip()
                        
                        adresse_entreprise_parts = [p for p in [numero_entreprise, rue_entreprise, code_postal_entreprise, ville_entreprise] if p]
                        adresse_entreprise = ', '.join(adresse_entreprise_parts) if adresse_entreprise_parts else None
                        
                        if adresse_entreprise:
                            entreprise.adresse = adresse_entreprise
                        
                        # Mettre à jour les champs de restauration et QPV
                        if specialite_culinaire:
                            entreprise.specialite_culinaire = specialite_culinaire
                        if nom_concept:
                            entreprise.nom_concept = nom_concept
                        if site_internet:
                            entreprise.site_internet = site_internet
                        if lien_reseaux_sociaux:
                            entreprise.lien_reseaux_sociaux = lien_reseaux_sociaux
                        entreprise.qpv = qpv
                        
                        # Géocoder l'adresse entreprise si elle existe et que les coordonnées ne sont pas déjà présentes
                        if adresse_entreprise and (not entreprise.lat or not entreprise.lng):
                            try:
                                latlng = await geocode_one(adresse_entreprise)
                                if latlng:
                                    entreprise.lat, entreprise.lng = latlng
                            except Exception as e:
                                logging.warning(f"Erreur lors du géocodage de l'adresse entreprise (mise à jour, ligne {idx + 2}): {e}")
                
                # Calculer et stocker l'éligibilité
                eligibilite_inserted = False
                try:
                    # Vérifier que la table eligibilite existe dans le schéma
                    if not table_exists_anywhere('eligibilite', session, schema_name):
                        logging.warning(f"Table eligibilite n'existe pas dans le schéma {schema_name}, création...")
                        schema_service._create_tables_in_schema(schema_name)
                        session.commit()
                    
                    # Récupérer l'entreprise pour l'ancienneté et l'adresse (toujours récupérer, qu'elle soit créée ou mise à jour)
                    entreprise_for_elig = None
                    adresse_entreprise_for_elig = None
                    
                    # Si on vient de créer l'entreprise, elle est déjà dans la variable 'entreprise'
                    if doit_creer_entreprise and 'entreprise' in locals() and entreprise:
                        entreprise_for_elig = entreprise
                    elif not doit_creer_entreprise and 'entreprise' in locals() and entreprise:
                        # Si l'entreprise existe déjà, elle a été récupérée dans le bloc else précédent
                        entreprise_for_elig = entreprise
                    else:
                        # Sinon, récupérer l'entreprise existante (par SIRET ou candidat_id)
                        if siret and siret != 'nan':
                            entreprise_for_elig = session.exec(select(EntrepriseSchema).where(EntrepriseSchema.siret == siret)).first()
                        if not entreprise_for_elig:
                            entreprise_for_elig = session.exec(select(EntrepriseSchema).where(EntrepriseSchema.candidat_id == cand.id)).first()
                    
                    if entreprise_for_elig:
                        adresse_entreprise_for_elig = entreprise_for_elig.adresse
                    
                    # Calculer l'ancienneté
                    anciennete = entreprise_age_annees(entreprise_for_elig.date_creation if entreprise_for_elig else None)
                    if anciennete is not None:
                        anciennete = int(anciennete)  # Convertir en entier pour correspondre au type attendu
                    
                    # Utiliser le chiffre d'affaires de l'entreprise si disponible, sinon celui de la préinscription
                    ca_for_elig = chiffre_affaires_preinscription
                    if entreprise_for_elig and entreprise_for_elig.chiffre_affaires:
                        ca_for_elig = str(entreprise_for_elig.chiffre_affaires)
                    
                    # Calculer l'éligibilité
                    verdict, details = await evaluate_eligibilite(
                        adresse_perso=adresse_personnelle,
                        adresse_entreprise=adresse_entreprise_for_elig,
                        chiffre_affaires=ca_for_elig,
                        anciennete_annees=anciennete,
                        programme_id=prog.id,
                        session=session,
                        request=request,
                        preinscription_id=preinscription.id,
                        schema_name=schema_name
                    )
                    
                    # Mettre à jour QPV dans l'entreprise si disponible depuis evaluate_eligibilite
                    if entreprise_for_elig and details.get("qpv_ok"):
                        qpv_nom_from_details = details.get("qpv_ok")
                        if qpv_nom_from_details.startswith("QPV"):
                            entreprise_for_elig.qpv = True
                    
                    # L'éligibilité est maintenant enregistrée automatiquement par evaluate_eligibilite
                    # Plus besoin d'insertion manuelle
                    eligibilite_inserted = True
                    
                except Exception as e:
                    logging.error(f"Erreur lors de l'insertion de l'éligibilité (ligne {idx + 2}): {e}")
                    import traceback
                    logging.error(traceback.format_exc())
                    # Ne pas faire échouer l'import si l'éligibilité échoue
                    eligibilite_inserted = False
                
                # Commit final (tout est déjà dans le schéma du programme)
                session.commit()
                
                imported_count += 1
                logging.info(f"✅ Ligne {idx + 2} importée avec succès - Total: {imported_count}")
                
            except Exception as e:
                error_msg = f"Ligne {idx + 2}: {str(e)}"
                logging.error(f"Erreur lors de l'import de la ligne {idx + 2}: {error_msg}")
                import traceback
                logging.error(traceback.format_exc())
                errors.append(error_msg)
                session.rollback()
                continue
        
        # Message de résultat
        if imported_count > 0:
            message = f"{imported_count} préinscription(s) importée(s) avec succès"
            if errors:
                message += f". {len(errors)} erreur(s) rencontrée(s)"
            if warnings:
                message += f". {len(warnings)} avertissement(s)"
        else:
            if errors:
                message = f"Aucune préinscription importée. {len(errors)} erreur(s) rencontrée(s)"
            else:
                message = "Aucune préinscription à importer"
        
        logging.info(f"✅ Import Excel terminé: {imported_count} préinscription(s) importée(s), {len(errors)} erreur(s), {len(warnings)} avertissement(s)")
        
        # Rediriger vers la liste avec un message via query parameters (plus fiable que headers)
        from fastapi import status
        from urllib.parse import urlencode
        redirect_url = request.url_for("preinscriptions_form")
        if programme_code:
            redirect_url = f"{redirect_url}?programme={programme_code}"
        
        # Ajouter le message et les erreurs dans l'URL
        params = {
            "import_success": "true" if imported_count > 0 else "false",
            "imported": str(imported_count),
            "message": message
        }
        
        # Ajouter les erreurs si présentes
        if errors:
            params["errors_count"] = str(len(errors))
            # Ajouter les détails des erreurs (limitées aux 20 premières pour éviter une URL trop longue)
            errors_to_show = errors[:20]  # Limiter à 20 erreurs pour éviter une URL trop longue
            # Encoder le JSON en base64 pour éviter les problèmes d'encodage URL
            errors_json = json.dumps(errors_to_show, ensure_ascii=False)
            params["errors_details"] = base64.urlsafe_b64encode(errors_json.encode('utf-8')).decode('utf-8')
            if len(errors) > 20:
                params["errors_total"] = str(len(errors))  # Nombre total d'erreurs si on en a plus de 20
        
        # Ajouter les avertissements si présents
        if warnings:
            params["warnings_count"] = str(len(warnings))
            warnings_to_show = warnings[:10]  # Limiter à 10 avertissements
            warnings_json = json.dumps(warnings_to_show, ensure_ascii=False)
            params["warnings_details"] = base64.urlsafe_b64encode(warnings_json.encode('utf-8')).decode('utf-8')
        
        redirect_url_with_params = f"{redirect_url}&{urlencode(params)}"
        
        return RedirectResponse(
            url=redirect_url_with_params,
            status_code=status.HTTP_302_FOUND,
            headers={"X-Message": message, "X-Errors": json.dumps(errors) if errors else ""}
        )
        
    except HTTPException as http_exc:
        # Capturer les HTTPException et les rediriger avec un message
        logging.error(f"Erreur HTTP lors de l'import Excel: {http_exc.status_code} - {http_exc.detail}")
        
        # Rollback de la session
        try:
            session.rollback()
        except:
            pass
        
        # Rediriger vers la liste avec un message d'erreur
        from fastapi import status
        from urllib.parse import urlencode
        redirect_url = request.url_for("preinscriptions_form")
        if programme_code:
            redirect_url = f"{redirect_url}?programme={programme_code}"
        
        error_message = str(http_exc.detail) if http_exc.detail else f"Erreur HTTP {http_exc.status_code}"
        params = {
            "import_success": "false",
            "imported": "0",
            "message": error_message,
            "errors_count": "1",
            "error_type": "HTTPException"
        }
        
        # Ajouter les détails de l'erreur
        error_details = [error_message]
        errors_json = json.dumps(error_details, ensure_ascii=False)
        params["errors_details"] = base64.urlsafe_b64encode(errors_json.encode('utf-8')).decode('utf-8')
        
        redirect_url_with_params = f"{redirect_url}&{urlencode(params)}"
        
        return RedirectResponse(
            url=redirect_url_with_params,
            status_code=status.HTTP_302_FOUND
        )
    except Exception as e:
        import traceback
        error_traceback = traceback.format_exc()
        error_type = type(e).__name__
        logging.error(f"Erreur lors de l'import Excel ({error_type}): {e}")
        logging.error(error_traceback)
        
        # Rollback de la session en cas d'erreur globale
        try:
            session.rollback()
        except:
            pass
        
        # Construire un message d'erreur détaillé mais concis
        error_message = f"Erreur lors de l'import: {str(e)}"
        
        # Ajouter le type d'erreur pour aider au diagnostic
        error_type = type(e).__name__
        if error_type != "Exception":
            error_message = f"[{error_type}] {error_message}"
        
        # Au lieu de lever une HTTPException, rediriger vers la liste avec un message d'erreur
        # Cela permet au frontend d'afficher le message via les query parameters
        from fastapi import status
        from urllib.parse import urlencode
        redirect_url = request.url_for("preinscriptions_form")
        if programme_code:
            redirect_url = f"{redirect_url}?programme={programme_code}"
        
        params = {
            "import_success": "false",
            "imported": "0",
            "message": error_message,
            "errors_count": "1",
            "error_type": error_type
        }
        
        # Ajouter les détails de l'erreur si ce n'est pas trop long
        if len(error_message) < 200:
            error_details = [error_message]
            errors_json = json.dumps(error_details, ensure_ascii=False)
            params["errors_details"] = base64.urlsafe_b64encode(errors_json.encode('utf-8')).decode('utf-8')
        
        redirect_url_with_params = f"{redirect_url}&{urlencode(params)}"
        
        return RedirectResponse(
            url=redirect_url_with_params,
            status_code=status.HTTP_302_FOUND
        )